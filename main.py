# main.py
# 导入 os 库
import os
# 导入 pandas 库
import pandas as pd
# 导入 openpyxl 库
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore")
import psycopg2


def test_db_connection(host, user, password, database, port):
    # 尝试连接数据库
    try:
        # 创建一个连接对象
        connection = psycopg2.connect(host=host,
                                      user=user,
                                      password=password,
                                      database=database,
                                      port=port)
        # 如果没有抛出异常，说明连接成功
        print(f'Successfully connected to {database}')  # 打印日志
        return True
    # 如果抛出异常，说明连接失败
    except Exception as e:
        # 打印异常信息
        print(e)
        # 返回 False
        return False


# 定义一个函数，用于执行 sql 文件并保存到 excel 表格中，接收六个参数：数据库的链接信息，.sql 文件的路径，excel 输出的路径
def execute_sql_to_excel(host, user, password, database, sql_path, excel_path, port):
    # 连接数据库
    try:
        conn = psycopg2.connect(host=host, user=user,
                                password=password,
                                database=database,
                                charset='utf8',
                                port=port)
        cursor = conn.cursor()
        print(f'Connected to {database} with cursor')  # 打印日志
    except Exception as e:
        print(f'Error connecting to database: {e}')
        return

    # 获取 sql 文件夹中的所有 .sql 文件
    sql_files = [f for f in os.listdir(sql_path) if f.endswith('.sql')]
    print(f'Found {len(sql_files)} sql files in {sql_path}')  # 打印日志
    # 执行所有 sql 文件
    for file in sql_files:
        # 打开文件
        f = open(os.path.join(sql_path,
                              file), 'r', encoding='utf8')
        sql = f.read()
        # 分割 sql 为多个语句
        statements = sql.split(';')
        print(f'Read {len(statements)} statements from {file}')  # 打印日志
        for statement in statements:
            # 去掉空白语句
            if statement.strip():
                try:
                    # 重新获取游标
                    cursor = conn.cursor()
                    cursor.execute(statement)
                    results = cursor.fetchall()
                    # 获取第一列的值并去重排序
                    values = sorted(set([row[0] for row in results if row[0]]))
                    print(f'Found {len(values)} distinct values in the first column')  # 打印日志
                    # 根据去重后的值判断输入到哪个表格中
                    for value in values:
                        # 筛选出第一列等于 value 的结果
                        filtered_results = [row for row in results if row[0] == value]
                        df = pd.DataFrame(filtered_results, columns=[i[0] for i in cursor.description])
                        # 拼接 excel 输出的完整路径
                        excel_file = os.path.join(excel_path, f'{value}.xlsx')
                        if os.path.exists(excel_file):
                            book = load_workbook(excel_file)
                            writer = pd.ExcelWriter(excel_file, engine='openpyxl')
                            writer.book = book
                            # 判断是否已经存在同名的 sheet
                            if os.path.splitext(file)[0] in book.sheetnames:
                                print(f'Warning: sheet {os.path.splitext(file)[0]} already exists in {value}.xlsx')
                            else:
                                # 修改 header 参数为 True
                                df.to_excel(writer, index=False, header=True, startrow=0,
                                            sheet_name=os.path.splitext(file)[0])
                                print(
                                    f'Wrote {len(df)} rows to sheet {os.path.splitext(file)[0]} in {value}.xlsx')  # 打印日志

                            try:
                                writer.save()
                            except Exception as e:
                                print(f'Error saving {value}.xlsx: {e}')
                        else:
                            df.to_excel(excel_file, index=False, sheet_name=os.path.splitext(file)[0])
                except Exception as e:
                    print(f'Error executing {file}: {e}')

        # 关闭文件
        f.close()

    # 关闭数据库连接
    if not conn._closed:  # 判断数据库连接是否已经关闭
        cursor.close()
        conn.close()


