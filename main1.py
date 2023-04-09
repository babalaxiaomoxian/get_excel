# main.py
# 导入 os 库
import os
# 导入 pandas 库
import pandas as pd
# 导入 openpyxl 库
from openpyxl import load_workbook
import warnings

warnings.filterwarnings("ignore")
import psycopg2
# 导入 tqdm 库
from tqdm import tqdm

def test_db_connection(host, user, password, database, port):
    # 尝试连接数据库
    try:
        # 创建一个连接对象
        connection = psycopg2.connect(host=host,
                                      user=user,
                                      password=password,
                                      database=database,
                                      port=port)
        # 如果没有抛出异常，说明连接成功
        print(f'Successfully connected to {database}')  # 打印日志
        return True
    # 如果抛出异常，说明连接失败
    except Exception as e:
        # 打印异常信息
        print(e)
        # 返回 False
        return False


def execute_sql_to_excel(host, user, password, database, sql_path, excel_path, port):
    # 连接数据库
    try:
        conn = psycopg2.connect(host=host, user=user,
                                password=password,
                                database=database,
                                # charset='utf8',
                                port=port)
        cursor = conn.cursor()
        print(f'Connected to {database} with cursor')  # 打印日志
    except Exception as e:
        print(f'Error connecting to database: {e}')
        return

    # 获取 sql 文件夹中的所有 .sql 文件
    sql_files = [f for f in os.listdir(sql_path) if f.endswith('.sql')]
    print(f'Found {len(sql_files)} sql files in {sql_path}')  # 打印日志

    # 执行所有 sql 文件
    for file in sql_files:
        # 打开文件
        f = open(os.path.join(sql_path,
                              file), 'r', encoding='utf8')
        sql = f.read()
        # 分割 sql 为多个语句
        statements = sql.split(';')
        print(f'Read {len(statements)} statements from {file}')  # 打印日志
        for statement in statements:
            # 去掉空白语句
            if statement.strip():
                try:
                    # 重新获取游标
                    cursor = conn.cursor()
                    cursor.execute(statement)
                    results = cursor.fetchall()
                    # 获取第一列的值并去重排序
                    values = sorted(set([row[0] for row in results if row[0]]))
                    print(f'Found {len(values)} distinct values in the first column')  # 打印日志
                    # 根据去重后的值判断输入到哪个表格中
                    for value in values:
                        # 筛选出第一列等于 value 的结果
                        filtered_results = [row for row in results if row[0] == value]
                        df = pd.DataFrame(filtered_results, columns=[i[0] for i in cursor.description])
                        # 拼接 excel 输出的完整路径
                        excel_file = os.path.join(excel_path, f'{value}.xlsx')
                        if os.path.exists(excel_file):
                            book = load_workbook(excel_file)
                            writer = pd.ExcelWriter(excel_file, engine='openpyxl')
                            writer.book = book
                            # 判断是否已经存在同名的 sheet
                            if os.path.splitext(file)[0] in book.sheetnames:
                                print(f'Warning: sheet {os.path.splitext(file)[0]} already exists in {value}.xlsx')
                            else:
                                # 计算当前的 sheet 页有多少行数据要写入
                                total_rows = len(df)
                                print(f'Found {total_rows} rows to write to sheet {os.path.splitext(file)[0]} in {value}.xlsx')  # 打印日志

                                # 初始化一个变量，用于记录当前写入了多少行数据
                                current_rows = 0

                                # 创建一个 tqdm 对象，用于显示进度条
                                pbar = tqdm(total=total_rows)

                                # 定义一个变量，用于记录每次写入的行数
                                chunksize = 1000

                                # 用 for 循环来实现每次写入 chunksize 行数据
                                for i in range(0, total_rows, chunksize):
                                    # 取出第 i 到 i + chunksize 行数据
                                    chunk_df = df.iloc[i:i + chunksize]
                                    # 修改 header
                                    # 修改 header 参数为 True，如果是第一次写入，否则为 False
                                    header = True if i == 0 else False
                                    # 写入到 excel 中
                                    chunk_df.to_excel(writer, index=False, header=header, startrow=i,
                                            sheet_name=os.path.splitext(file)[0])
                                    print(
                                        f'Wrote {len(chunk_df)} rows to sheet {os.path.splitext(file)[0]} in {value}.xlsx')  # 打印日志

                                    # 更新当前写入的行数
                                    current_rows += len(chunk_df)

                                    # 更新进度条
                                    pbar.update(current_rows)

                                # 关闭进度条
                                pbar.close()

                            try:
                                writer.save()
                            except Exception as e:
                                print(f'Error saving {value}.xlsx: {e}')
                        else:
                            df.to_excel(excel_file, index=False, sheet_name=os.path.splitext(file)[0])
                except Exception as e:
                    print(f'Error executing {file}: {e}')

        # 关闭文件
        f.close()

    # 关闭数据库连接
    if not conn.closed:  # 判断数据库连接是否已经关闭
        cursor.close()
        conn.close()
